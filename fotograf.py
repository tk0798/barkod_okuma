from pyzbar.pyzbar import decode
from PIL import Image
from openpyxl import Workbook,load_workbook
import os
i = 0

wb = Workbook()

ws = wb.active
ws.title = "İlk Çalışma Alanı"

print(wb.sheetnames)
dosyalar = os.listdir("C:\\Users\\kaant\\PycharmProjects\\barkod_qr\\fotolar")

for dosya in dosyalar:
    konum = "C:\\Users\\kaant\\PycharmProjects\\barkod_qr\\fotolar\\"+dosya
    for bar in decode(Image.open(konum)):
        i = i + 1
        print(i)
        print(bar.data)
        ws.append([bar.data,konum])

    wb.save("C:\\Users\\kaant\\PycharmProjects\\barkod_qr\\dosya.xlsx")
input()

