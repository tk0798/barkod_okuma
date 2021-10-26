from pyzbar.pyzbar import decode
from PIL import Image
import os
import requests
import cv2
import numpy as np
import imutils
from pyzbar import pyzbar
import time
from openpyxl import Workbook,load_workbook
import yollar

print("Fotograf icin 1 Video icin 2 cıkıs yapmak icin ise herhangi bir tusa basınız :")
x = input()


if x == "1":

    print("kaydedilecek excel dosyasının adını giriniz :")
    fotograf_cikti_excel = input()
    say = 0
    i = 0

    try:

        wb = load_workbook(yollar.barkod_fotograf_cikti_excel +"\\" + fotograf_cikti_excel + ".xlsx")
        say = 1

    except:
        wb = Workbook()

    ws = wb.active
    print("Kaydedilecek sayfa adını giriniz :")

    ws.title = input()

    # if say == 1:
    #     wb.create_sheet(ws.title,i)


    print(wb.sheetnames)
    dosyalar = os.listdir(yollar.barkod_okutulacak_fotograf_yolu)

    for dosya in dosyalar:
        konum = yollar.barkod_okutulacak_fotograf_yolu + "\\" + dosya
        for bar in decode(Image.open(konum)):
            i = i + 1
            print(i)
            print(bar.data)
            ws.append([bar.data, konum])

        wb.save(yollar.barkod_fotograf_cikti_excel +"\\" + fotograf_cikti_excel + ".xlsx")
    input()
elif x == "2":

    print("Kaydedilecek excel dosyasının adını giriniz :")
    video_cikti_excel = input()

    try:

        wb = load_workbook(yollar.barkod_video_cikti_excel+"\\" + video_cikti_excel + ".xlsx")


    except:
        wb = Workbook()

    ws = wb.active
    ws.title = "İlk Çalışma Alanı"
    # Replace the below URL with your own. Make sure to add "/shot.jpg" at last.
    url = yollar.ipcam_url + "/shot.jpg"
    found = set()
    data = []
    # While loop to continuously fetching data from the Url
    say = 0
    while True:
        img_resp = requests.get(url)
        img_arr = np.array(bytearray(img_resp.content), dtype=np.uint8)
        img = cv2.imdecode(img_arr, -1)
        img = imutils.resize(img, width=640, height=480)

        # find the barcodes in the frame and decode each of the barcodes
        barcodes = pyzbar.decode(img)

        # loop over the detected barcodes
        for num, barcode in enumerate(barcodes):
            # extract the bounding box location of the barcode and draw
            # the bounding box surrounding the barcode on the image
            barcodeData = barcode.data.decode("utf-8")
            barcodeType = barcode.type

            if barcodeData not in found and barcodeType == "CODE128":
                # if len(barcodeData) == 14:
                (x, y, w, h) = barcode.rect
                cv2.rectangle(img, (x, y), (x + w, y + h), (0, 0, 255), 2)

                # the barcode data is a bytes object so if we want to draw it
                # on our output image we need to convert it to a string first

                # draw the barcode data and barcode type on the image
                text = "{} ({})".format(barcodeData, barcodeType)
                cv2.putText(img, text, (x, y - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)


            elif barcodeType == "CODE128":
                # if len(barcodeData) == 14:
                (x, y, w, h) = barcode.rect
                cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)

                # the barcode data is a bytes object so if we want to draw it
                # on our output image we need to convert it to a string first

                # draw the barcode data and barcode type on the image
                text = "{} ({})".format(barcodeData, barcodeType)
                cv2.putText(img, text, (x, y - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
            if barcodeData not in found and barcodeType == "CODE128":
                # if len(barcodeData) == 14:
                found.add(barcodeData)
                print(barcodeData)
                data.append(barcodeData)
                ws.append([barcodeData])
                say = say + 1


        # if the barcode text is currently not in our CSV file, write
        # the timestamp + barcode to disk and update the set

        # show the output frame
        cv2.imshow("Barcode Scanner", img)
        key = cv2.waitKey(1) & 0xFF


        k = cv2.waitKey(10) & 0xff
        if k == 27:
            break

    # close the output CSV file do a bit of cleanup
    print("[INFO] cleaning up...")

    cv2.destroyAllWindows()
    print(data)
    print(len(data))
    wb.save(yollar.barkod_video_cikti_excel+"\\" + video_cikti_excel + ".xlsx")
elif x != 1 and x != 2:
    exit()